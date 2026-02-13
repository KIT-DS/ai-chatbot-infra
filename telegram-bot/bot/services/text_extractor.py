import csv
import io
from pathlib import Path

from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook


class UnsupportedFormatError(Exception):
    pass


def extract_text(file_path: str, file_name: str) -> str:
    """Extract text from a file based on its extension."""
    ext = Path(file_name).suffix.lower()
    extractors = {
        ".pdf": _extract_pdf,
        ".docx": _extract_docx,
        ".txt": _extract_txt,
        ".xlsx": _extract_xlsx,
        ".csv": _extract_csv,
    }

    extractor = extractors.get(ext)
    if extractor is None:
        raise UnsupportedFormatError(f"Неподдерживаемый формат файла: {ext}")

    text = extractor(file_path)
    if not text or not text.strip():
        raise ValueError("Файл пуст или не содержит извлекаемого текста.")

    return text.strip()


def _extract_pdf(path: str) -> str:
    reader = PdfReader(path)
    pages = []
    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            pages.append(page_text)
    return "".join(pages)


def _extract_docx(path: str) -> str:
    doc = Document(path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "".join(paragraphs)


def _extract_txt(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def _extract_xlsx(path: str) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"--- Sheet: {sheet_name} ---" + "".join(rows))
    wb.close()
    return "".join(parts)


def _extract_csv(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        rows = []
        for row in reader:
            if any(cell.strip() for cell in row):
                rows.append("\t".join(row))
    return "".join(rows)
